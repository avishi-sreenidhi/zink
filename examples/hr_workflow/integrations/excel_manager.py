"""
ExcelManager: Manages HR pipeline spreadsheet
Logs candidate screening decisions to Excel
"""

from openpyxl import Workbook, load_workbook
from pathlib import Path
from datetime import datetime


class ExcelManager:
    """
    Manages hr_pipeline.xlsx spreadsheet
    
    Columns: Candidate ID | Name | Email | Status | Screening Score | Decision | Timestamp
    """
    
    def __init__(self, filename: str = "hr_pipeline.xlsx"):
        self.filename = Path(filename)
        self.ensure_file_exists()
    
    def ensure_file_exists(self):
        """Create Excel file if it doesn't exist"""
        if not self.filename.exists():
            wb = Workbook()
            ws = wb.active
            ws.title = "Pipeline"
            
            # Headers
            headers = [
                "Candidate ID",
                "Name",
                "Email",
                "Status",
                "Screening Score",
                "Decision",
                "Timestamp"
            ]
            ws.append(headers)
            
            wb.save(self.filename)
            print(f"✓ Created Excel file: {self.filename}")
    
    def add_candidate(self, candidate_id: str, name: str, email: str) -> None:
        """Add new candidate to pipeline"""
        wb = load_workbook(self.filename)
        ws = wb.active
        
        # Append new row
        ws.append([
            candidate_id,
            name,
            email,
            "screening",           # Status
            None,                   # Score (will be updated later)
            None,                   # Decision (will be updated later)
            datetime.now().isoformat()
        ])
        
        wb.save(self.filename)
        print(f"✓ Added candidate: {name} ({candidate_id})")
    
    def update_screening_score(self, candidate_id: str, score: float, decision: str) -> None:
        """Update screening score and decision for candidate"""
        wb = load_workbook(self.filename)
        ws = wb.active
        
        # Find the row with this candidate_id
        for row in ws.iter_rows(min_row=2, values_only=False):  # Skip header
            if row[0].value == candidate_id:  # Column A = Candidate ID
                row[4].value = score           # Column E = Screening Score
                row[5].value = decision        # Column F = Decision
                row[3].value = f"screening_{decision}"  # Column D = Status
                break
        
        wb.save(self.filename)
        print(f"✓ Updated screening: {candidate_id} - Score: {score:.2f}, Decision: {decision}")
    
    def update_interview_date(self, candidate_id: str, interview_date: str) -> None:
        """Update interview schedule for candidate"""
        wb = load_workbook(self.filename)
        ws = wb.active
        
        # Find the row
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value == candidate_id:
                row[3].value = f"interview_scheduled_{interview_date}"
                break
        
        wb.save(self.filename)
        print(f"✓ Scheduled interview: {candidate_id} - {interview_date}")
    
    def update_approval(self, candidate_id: str, approval_decision: str, salary: float = None) -> None:
        """Update approval decision"""
        wb = load_workbook(self.filename)
        ws = wb.active
        
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value == candidate_id:
                row[3].value = f"approved_{approval_decision}"
                if salary:
                    row[6].value = f"Salary: ${salary:,.0f}"
                break
        
        wb.save(self.filename)
        print(f"✓ Updated approval: {candidate_id} - {approval_decision} (${salary:,.0f})" if salary else f"✓ Updated approval: {candidate_id} - {approval_decision}")
    
    def get_all_candidates(self) -> list[dict]:
        """Fetch all candidates from pipeline"""
        wb = load_workbook(self.filename)
        ws = wb.active
        
        candidates = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Only if candidate_id exists
                candidates.append({
                    "candidate_id": row[0],
                    "name": row[1],
                    "email": row[2],
                    "status": row[3],
                    "score": row[4],
                    "decision": row[5],
                    "timestamp": row[6]
                })
        
        return candidates